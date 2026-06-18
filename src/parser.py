import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import List, Optional, Tuple
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from src.config import SHEET_SOURCE

TZ_WIB = ZoneInfo("Asia/Jakarta")

@dataclass
class AccountRow:
    no: int
    nama_kanwil: str
    url_akun: str
    manual_judul: str
    manual_link: str
    manual_reach: str
    agenda_no: str
    agenda_topic: str

@dataclass
class ScrapeRow:
    nama_kanwil: str
    url_akun: str
    post_url: str
    shortcode: str
    tanggal_postingan: Optional[datetime]
    media_type: str
    caption: str
    like_count: Optional[int]
    comment_count: Optional[int]
    total_engagement: Optional[int]
    status_periode: str
    status_scraping: str
    catatan: str

def safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()

def normalize_url(url: str) -> str:
    url = safe_text(url)
    if not url:
        return ""
    if not url.startswith("http"):
        url = "https://" + url
    return url.strip()

def extract_shortcode(url: str) -> str:
    url = normalize_url(url)
    if not url:
        return ""
    try:
        parsed = urlparse(url)
        parts = [p for p in parsed.path.split("/") if p]
        if len(parts) >= 2 and parts[0] in ["p", "reel", "tv"]:
            return parts[1]
        if len(parts) >= 3 and parts[1] in ["p", "reel", "tv"]:
            return parts[2]
        return ""
    except Exception:
        return ""

def detect_media_type(post_url: str) -> str:
    if "/reel/" in post_url:
        return "Reels"
    if "/p/" in post_url:
        return "Post / Picture / Carousel"
    if "/tv/" in post_url:
        return "Video"
    return "Unknown"

def clean_caption(text: str, max_len: int = 800) -> str:
    text = safe_text(text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return ""
    return text[:max_len]

def parse_dt_to_wib(raw: str) -> Optional[datetime]:
    raw = safe_text(raw)
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.astimezone(TZ_WIB).replace(tzinfo=None)
    except Exception:
        return None

def parse_number_token(token: str) -> Optional[int]:
    token = safe_text(token).lower()
    if not token:
        return None
    token = token.replace(" ", "")
    multiplier = 1
    suffix_map = {
        "k": 1_000,
        "rb": 1_000,
        "ribu": 1_000,
        "m": 1_000_000,
        "jt": 1_000_000,
        "juta": 1_000_000,
    }
    for suffix, value in suffix_map.items():
        if token.endswith(suffix):
            multiplier = value
            token = token[: -len(suffix)]
            break
    token = token.strip()
    if not token:
        return None
    if multiplier > 1:
        token = token.replace(",", ".")
        try:
            return int(float(token) * multiplier)
        except Exception:
            return None
    if "," in token and "." in token:
        token = token.replace(",", "")
    elif "," in token:
        before, after = token.split(",", 1)
        if len(after) == 3:
            token = before + after
        else:
            token = token.replace(",", ".")
    elif "." in token:
        before, after = token.split(".", 1)
        if len(after) == 3:
            token = before + after
    try:
        return int(float(token))
    except Exception:
        return None

def parse_engagement_from_text(text: str) -> Tuple[Optional[int], Optional[int]]:
    text = safe_text(text)
    if not text:
        return None, None
    normalized = text.lower()
    number_pattern = r"([\d]+(?:[.,][\d]+)?(?:\s*(?:k|m|rb|ribu|jt|juta))?)"
    like_patterns = [
        rf"{number_pattern}\s+likes?",
        rf"{number_pattern}\s+suka",
    ]
    comment_patterns = [
        rf"{number_pattern}\s+comments?",
        rf"{number_pattern}\s+komentar",
    ]
    like_count = None
    comment_count = None
    for pattern in like_patterns:
        match = re.search(pattern, normalized)
        if match:
            like_count = parse_number_token(match.group(1))
            break
    for pattern in comment_patterns:
        match = re.search(pattern, normalized)
        if match:
            comment_count = parse_number_token(match.group(1))
            break
    return like_count, comment_count

def extract_caption_from_meta(*meta_values: str) -> str:
    combined = ""
    for value in meta_values:
        value = safe_text(value)
        if value:
            combined = value
            break
    if not combined:
        return ""
    text = re.sub(r"\s+", " ", combined).strip()
    if ": " in text:
        possible_caption = text.split(": ", 1)[-1]
        possible_caption = possible_caption.strip().strip('"').strip("“").strip("”")
        if possible_caption:
            return clean_caption(possible_caption)
    if " - " in text:
        possible_caption = text.split(" - ", 1)[-1]
        possible_caption = possible_caption.strip().strip('"').strip("“").strip("”")
        if possible_caption:
            return clean_caption(possible_caption)
    return clean_caption(text)

def status_periode(tanggal: Optional[datetime], period_start: datetime, period_end: datetime) -> str:
    if tanggal is None:
        return "Perlu Cek Manual"
    if period_start <= tanggal <= period_end:
        return "Masuk Periode"
    return "Di Luar Periode"

def load_accounts_from_excel(file_bytes: bytes) -> List[AccountRow]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    sheet = None

    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == SHEET_SOURCE.lower():
            sheet = workbook[sheet_name]
            break

    if sheet is None:
        for candidate_sheet in workbook.worksheets:
            found_instagram = False

            for row in candidate_sheet.iter_rows(values_only=True):
                for value in row:
                    if "instagram.com" in safe_text(value).lower():
                        found_instagram = True
                        break

                if found_instagram:
                    break

            if found_instagram:
                sheet = candidate_sheet
                break

    if sheet is None:
        raise ValueError("Tidak ada sheet yang berisi link Instagram. Pastikan file berisi URL instagram.com.")

    def to_profile_url(value: str) -> str:
        value = normalize_url(value)

        if not value:
            return ""

        try:
            parsed = urlparse(value)
            parts = [part for part in parsed.path.split("/") if part]

            if not parts:
                return ""

            if parts[0] in ["p", "reel", "tv", "stories"]:
                return ""

            username = parts[0].strip().lower()

            if not username:
                return ""

            return f"https://www.instagram.com/{username}/"

        except Exception:
            return ""

    def username_from_url(value: str) -> str:
        profile_url = to_profile_url(value)

        if not profile_url:
            return ""

        parsed = urlparse(profile_url)
        parts = [part for part in parsed.path.split("/") if part]

        if not parts:
            return ""

        return parts[0].strip()

    accounts = []
    seen_key = set()

    for row_index in range(1, sheet.max_row + 1):
        row_values = [
            safe_text(sheet.cell(row=row_index, column=col_index).value)
            for col_index in range(1, sheet.max_column + 1)
        ]

        instagram_columns = []

        for col_index, value in enumerate(row_values, start=1):
            if "instagram.com" in value.lower():
                instagram_columns.append(col_index)

        if not instagram_columns:
            continue

        for url_col in instagram_columns:
            raw_url = safe_text(sheet.cell(row=row_index, column=url_col).value)
            profile_url = to_profile_url(raw_url)

            if not profile_url:
                continue

            nama = ""

            if url_col > 1:
                nama = safe_text(sheet.cell(row=row_index, column=url_col - 1).value)

            if "kanwil" not in nama.lower() or "djpb" not in nama.lower():
                nama = safe_text(sheet.cell(row=row_index, column=2).value)

            if "kanwil" not in nama.lower() or "djpb" not in nama.lower():
                for value in row_values:
                    value_lower = value.lower()

                    if "kanwil" in value_lower and "djpb" in value_lower:
                        nama = value
                        break

            if not nama:
                username = username_from_url(profile_url)
                nama = f"Kanwil DJPb {username}"

            key = profile_url.lower()

            if key in seen_key:
                continue

            seen_key.add(key)

            accounts.append(
                AccountRow(
                    no=len(accounts) + 1,
                    nama_kanwil=nama,
                    url_akun=profile_url,
                    manual_judul=safe_text(sheet.cell(row=row_index, column=6).value),
                    manual_link=normalize_url(sheet.cell(row=row_index, column=7).value),
                    manual_reach=safe_text(sheet.cell(row=row_index, column=9).value),
                    agenda_no=safe_text(sheet.cell(row=row_index, column=10).value),
                    agenda_topic=safe_text(sheet.cell(row=row_index, column=11).value),
                )
            )

    if not accounts:
        raise ValueError("Tidak ada akun Instagram Kanwil DJPb yang terbaca. Pastikan file memiliki nama Kanwil dan URL instagram.com.")

    return accounts